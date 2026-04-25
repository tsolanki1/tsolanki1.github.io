from __future__ import annotations

import json
import re
import sys
from datetime import date, datetime, time
from pathlib import Path

from openpyxl import load_workbook


def clean(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_phone(value):
    text = clean(value).replace(" ", "")
    digits = "".join(ch for ch in text if ch.isdigit())
    if len(digits) == 10:
        return f"{digits[:3]}-{digits[3:6]}-{digits[6:]}"
    return text


def extract_flight_ident(value):
    text = clean(value)
    if not text or text == "N/A":
        return ""

    direct_match = re.search(r"\b([A-Za-z0-9]{2,3})\s*-?\s*(\d{1,4})\b", text)
    if direct_match:
        return f"{direct_match.group(1)}{direct_match.group(2)}".upper()

    normalized = text.lower()
    airline_codes = {
        "american airlines": "AA",
        "american": "AA",
        "delta air lines": "DL",
        "delta": "DL",
        "united airlines": "UA",
        "united": "UA",
        "southwest airlines": "WN",
        "southwest": "WN",
        "sw": "WN",
        "spirit airlines": "NK",
        "spirit": "NK",
        "frontier airlines": "F9",
        "frontier": "F9",
    }

    airline_code = ""
    for prefix, code in airline_codes.items():
        if normalized.startswith(prefix):
            airline_code = code
            break

    flight_number = re.search(r"(\d{1,4})", normalized)
    if airline_code and flight_number:
        return f"{airline_code}{flight_number.group(1)}"

    return ""


def short_airport(value):
    text = clean(value)
    return text.split(" - ")[0] if " - " in text else text


def airport_need(value):
    text = clean(value)
    if text == "Yes":
        return "Needs pickup / dropoff"
    if text == "No":
        return "No airport transfer needed"
    return text or "Not listed"


def transport_need(value):
    text = clean(value)
    if text == "Yes":
        return "Needs hotel to mandir transport"
    if text == "No":
        return "No hotel shuttle needed"
    return text or "Not listed"


def hotel_from_accommodation(value):
    text = clean(value).lower()
    if "hotel" in text:
        return "Hotel requested"
    if text:
        return "Self-arranged"
    return ""


def driver_group(row):
    if clean(row.get("Do you need Airport Pick-up/Drop-off? ")) != "Yes":
        return "Self-arranged"
    parts = [
        short_airport(row.get("Arrival Airport")),
        clean(row.get("Arrival Date")),
        clean(row.get("Arrival Time")),
    ]
    return " | ".join(part for part in parts if part)


def build_notes(row):
    parts = []
    accommodation = clean(row.get("Accommodation"))
    pickup = clean(row.get("Airport Pickup Contact"))
    dropoff = clean(row.get("Airport Dropoff Contact"))
    if accommodation:
        parts.append(f"Stay: {accommodation}")
    if pickup:
        parts.append(f"Pickup contact: {pickup}")
    if dropoff:
        parts.append(f"Dropoff contact: {dropoff}")
    return " | ".join(parts)


def map_row(index, row):
    dietary = clean(row.get("Dietary restrictions"))
    accommodation = clean(row.get("Accommodation"))
    return {
        "id": index,
        "name": clean(row.get("Full Name")),
        "city": "",
        "role": clean(row.get("Wing")) or "Attendee",
        "phone": normalize_phone(row.get("Mobile")),
        "email": clean(row.get("Email Address")),
        "bapsEmail": clean(row.get("BAPS Email")),
        "arrivalAirport": short_airport(row.get("Arrival Airport")),
        "arrivalDate": clean(row.get("Arrival Date")),
        "arrivalTime": clean(row.get("Arrival Time")),
        "arrivalFlight": clean(row.get("Arrival Airline & Flight Number")),
        "arrivalFlightIdent": extract_flight_ident(row.get("Arrival Airline & Flight Number")),
        "departureAirport": short_airport(row.get("Departure Airport")),
        "departureDate": clean(row.get("Departure Date")),
        "departureTime": clean(row.get("Departure Time")),
        "departureFlight": clean(row.get("Departure Airline & Flight Number")),
        "departureFlightIdent": extract_flight_ident(row.get("Departure Airline & Flight Number")),
        "airportNeed": airport_need(row.get("Do you need Airport Pick-up/Drop-off? ")),
        "driverGroup": driver_group(row),
        "mealPreference": dietary or "Standard vegetarian",
        "foodNotes": dietary,
        "transportationNeed": transport_need(row.get("Do you need transport from Hotel to Mandir?")),
        "hotel": clean(row.get("Hotel")) or hotel_from_accommodation(accommodation),
        "rooming": clean(row.get("Hotel Confirmation")),
        "status": "Confirmed" if clean(row.get("Will you be attending?")) == "Yes" else clean(row.get("Will you be attending?")) or "Unknown",
        "accommodation": accommodation,
        "airportPickupContact": clean(row.get("Airport Pickup Contact")),
        "airportDropoffContact": clean(row.get("Airport Dropoff Contact")),
        "dinnerThursday": clean(row.get("Dinner (Thursday)")),
        "dinnerSaturday": clean(row.get("Dinner (Saturday)")),
        "breakfastSunday": clean(row.get("Breakfast (Sunday)")),
        "notes": build_notes(row),
    }


def workbook_to_event_data(source_path):
    workbook = load_workbook(source_path, data_only=True)
    sheet = workbook.worksheets[0]
    rows = list(sheet.iter_rows(values_only=True))
    headers = list(rows[0])
    records = [dict(zip(headers, row)) for row in rows[1:] if any(value not in (None, "") for value in row)]
    attendees = [map_row(index, row) for index, row in enumerate(records, start=1) if clean(row.get("Full Name"))]

    return {
        "event": {
            "name": "BAPS NA Communications MidYear Meeting",
            "location": "Dallas, Texas",
            "dates": "Update with meeting dates",
            "notes": [
                f"Imported from {Path(source_path).name}.",
                "The dashboard keeps all responses, and logistics teams can filter to confirmed attendees.",
            ],
        },
        "attendees": attendees,
    }


def main():
    if len(sys.argv) < 2:
        raise SystemExit("Usage: python sync_registration.py <input.xlsx> [output.js]")

    source = Path(sys.argv[1])
    output = Path(sys.argv[2]) if len(sys.argv) > 2 else Path(__file__).with_name("data.js")
    event_data = workbook_to_event_data(source)
    payload = "window.EVENT_DATA = " + json.dumps(event_data, indent=2) + ";\n"
    output.write_text(payload, encoding="utf-8")
    print(f"Wrote {output}")


if __name__ == "__main__":
    main()
